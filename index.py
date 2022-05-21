import os
import openpyxl
import requests
from bs4 import BeautifulSoup

URL = "https://ar.wikipedia.org/wiki/%D9%82%D8%A7%D8%A6%D9%85%D8%A9_%D8%A3%D9%81%D8%B6%D9%84_%D9%85%D8%A6%D8%A9_%D8%B1%D9%88%D8%A7%D9%8A%D8%A9_%D8%B9%D8%B1%D8%A8%D9%8A%D8%A9"
res = requests.get(url=URL).text
soup = BeautifulSoup(res, "html.parser")
novels = {}
wb = openpyxl.Workbook()  # open workboopk
ws = wb.active  # open sheet
result = soup.find("table", class_="wikitable").find_all("tr")[1::1]  # type: ignore
for row, items in enumerate(result):
    data = items.find_all(["th", "td"])
    links = items.find_all("a")
    for col in range(len(data)):
        try:
            if col <= len(links):
                ws.cell(row+1, col+1, str(data[col].text)).hyperlink = "https://ar.wikipedia.org/wiki/"+links[col-1].text
            else:
                ws.cell(row+1, col+1, str(data[col].text))
        except IndexError:pass
wb.save('test.xls')
# for items in soup.find('table', class_='wikitable').find_all('tr')[1::1]:  # type: ignore
#     data = items.find_all(['th','td'])
#     links = items.find_all('a')
#     try:
#         novels['index'] = data[0].text
#         novels['title'] = data[1].text
#         novels['auther'] = data[2].text
#         novels['country'] = data[3].text
#     except IndexError:pass
