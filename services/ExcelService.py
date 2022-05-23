import openpyxl
from services.SheetInterface import SheetService
class ExcelService(SheetService):
    def __init__(self, path) -> None:
            self.path = path
            try:
                self.wb = openpyxl.load_workbook(self.path)  
            except FileNotFoundError:
                self.wb = openpyxl.Workbook()# new workboopk
            self.ws = self.wb.active  # new sheet
        
    def write(self, row: int, col: int, data: str, hyperlink=None):
        if self.ws:
            self.ws.cell(row, col, data).hyperlink=hyperlink

    def save(self):
        if self.wb:
            self.wb.save(self.path)

    def get_max_row(self) -> int:
        if self.ws:
            return self.ws.max_row
        return 0

    def delete(self, row: int):
        if self.ws:
            for index in range(self.ws.max_column):
                self.ws.cell(row, index+1).value = None
    def read(self) -> dict:
        books = {}
        # print(self.ws.max_row)
        for r in range(self.ws.max_row):
            # print(self.ws.cell(r+1,1).value)
            # if self.ws.cell(r+1,1).value:
            books[str(self.ws.cell(r + 1, 1).value)] = {
                "id":self.ws.cell(r + 1, 1).value,
                "title":self.ws.cell(r + 1, 2).value,
                "author":self.ws.cell(r + 1, 3).value,
                "country":self.ws.cell(r + 1, 4).value,
            }
        return books

    def close(self):
        self.wb.close()
    
    def write_multiple(self, data: list):
        for index, row in enumerate(data):
            for col, value in enumerate(row):
                if isinstance(value, tuple):
                    self.write(index + 1, col +1, data=value[0], hyperlink=value[1])
                else:
                    self.write(index + 1, col +1, data=value)
        # with pd.ExcelWriter(self.path, engine='openpyxl') as writer:
        #     dataframe.to_excel(writer, 'sheet1')
    def __str__(self) -> str:
        return 'Excel'